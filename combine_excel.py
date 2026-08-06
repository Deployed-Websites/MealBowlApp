"""
combine_excel.py
~~~~~~~~~~~~~~~~
Reads a specific cell range from every .xlsx/.xls file in a folder,
combines them into a single output file (one source file = one column).
Re-runs are incremental — only new files are added.

Can be run in two ways:
    Interactive mode (double-click exe or run with no arguments):
        python combine_excel.py

    Command line mode:
        python combine_excel.py <folder> <output_file> <start_cell> <end_cell> [options]

Positional arguments
    folder        Path to folder containing source Excel files
    output_file   Path to the combined output .xlsx (created or updated)
    start_cell    Top-left cell of the range to read  e.g. B3
    end_cell      Bottom-right cell of the range      e.g. B52

Optional arguments
    --sheet SHEET          Sheet name or 0-based index to read from each file.
                           If omitted and each file has one sheet, that is used.
    --output-start-row N   Row in the output file where the first header goes (default: 1)
    --output-start-col N   Column in the output file where the first header goes (default: 1,
                           or 2 if --label-col is set)
    --label-col            Write row labels into the first column of the output
    --col-width N          Set all column widths to N. If omitted, Excel decides.
"""

import argparse
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import column_index_from_string, get_column_letter


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean_path(raw: str) -> str:
    """Strip surrounding quotes and whitespace from a pasted path.
    Handles Windows Explorer paths (double quotes) and Mac paths (single quotes).
    """
    return raw.strip().strip('"').strip("'").strip()


def parse_cell(cell_str: str):
    """Return (col_index, row_index) from a cell reference like 'B3'."""
    m = re.match(r"([A-Za-z]+)(\d+)", cell_str.strip())
    if not m:
        raise ValueError(f"Cannot parse cell reference: {cell_str!r}")
    return column_index_from_string(m.group(1)), int(m.group(2))


def resolve_sheet(wb, sheet_ref, filepath):
    """Return the correct worksheet from wb given a name, index, or None."""
    if sheet_ref is None:
        if len(wb.sheetnames) == 1:
            return wb.worksheets[0]
        else:
            wb.close()
            raise ValueError(
                f"{filepath.name}: no sheet specified and multiple sheets found. "
                f"Available: {wb.sheetnames}"
            )
    elif isinstance(sheet_ref, int):
        return wb.worksheets[sheet_ref]
    else:
        if sheet_ref not in wb.sheetnames:
            wb.close()
            raise ValueError(
                f"{filepath.name}: sheet {sheet_ref!r} not found. "
                f"Available: {wb.sheetnames}"
            )
        return wb[sheet_ref]


def read_range(filepath: Path, sheet_ref, col_start, row_start, col_end, row_end):
    """Read cached values from a range in an Excel file. Returns a flat list."""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    sheet = resolve_sheet(wb, sheet_ref, filepath)
    values = [
        row[0] or 0
        for row in sheet.iter_rows(min_row=row_start, max_row=row_end,
                                   min_col=col_start, max_col=col_end,
                                   values_only=True)
    ]
    wb.close()
    return values


def read_header_cell(filepath: Path, sheet_ref, col, row):
    """
    Reads a specific cell from an input file to use as the column header
    in the combined output, instead of the filename.
    Used when the user selects header source option (2) in interactive mode.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    sheet = resolve_sheet(wb, sheet_ref, filepath)
    value = next(sheet.iter_rows(min_row=row, max_row=row,
                                 min_col=col, max_col=col,
                                 values_only=True))[0]
    wb.close()
    return value


def style_header(cell):
    cell.font = Font(name="Arial", bold=True)
    cell.alignment = Alignment(horizontal="center", wrap_text=True)


def style_label(cell):
    cell.font = Font(name="Arial", italic=True, color="404040")
    cell.alignment = Alignment(horizontal="left")


# ---------------------------------------------------------------------------
# Core logic
# ---------------------------------------------------------------------------

def discover_files(folder: Path, exclude: Path):
    return sorted(
        p for p in folder.iterdir()
        if p.is_file()
        and p.suffix.lower() in (".xlsx", ".xls", ".xlsm")
        and p.resolve() != exclude.resolve()
        and not p.name.startswith("~$")
    )
    
def read_output_headers(ws, output_start_row, output_start_col):
    """
    Read existing headers from the output file's header row using iter_rows.
    Returns a dict of {header_value: col_index}.
    Only includes columns that also have data in the row below the header —
    empty data columns are treated as unprocessed and will be rewritten.
    """
    if ws.max_column < output_start_col:
        return {}

    # Read header row
    header_row = next(ws.iter_rows(min_row=output_start_row, max_row=output_start_row,
                                   min_col=output_start_col, max_col=ws.max_column,
                                   values_only=False))

    # Read first data row (one below header)
    data_row = next(ws.iter_rows(min_row=output_start_row + 1, max_row=output_start_row + 1,
                                 min_col=output_start_col, max_col=ws.max_column,
                                 values_only=False))

    existing = {}
    for hdr_cell, dat_cell in zip(header_row, data_row):
        if hdr_cell.value is not None and dat_cell.value is not None:
            existing[hdr_cell.value] = hdr_cell.column

    return existing


def build_or_update(
    folder: Path,
    output_path: Path,
    sheet_ref,
    col_start, row_start,
    col_end, row_end,
    label_col: bool,
    output_start_row: int,
    output_start_col: int,
    col_width,
    header_cell_col=None,
    header_cell_row=None,
):
    source_files = discover_files(folder, exclude=output_path)
    if not source_files:
        print("No Excel files found in folder.")
        return

    # Load existing output or create fresh
    if output_path.exists():
        wb_out = openpyxl.load_workbook(output_path)
        ws = wb_out.active
        existing_headers = read_output_headers(ws, output_start_row, output_start_col)
        next_col = max(existing_headers.values()) + 1 if existing_headers else output_start_col
    else:
        wb_out = openpyxl.Workbook()
        ws = wb_out.active
        ws.title = "Combined"
        existing_headers = {}
        next_col = output_start_col

    # Detect new files — also catches files whose data was deleted (header present but no data)
    new_files = []
    for f in source_files:
        col_header = f.stem if header_cell_row is None else read_header_cell(f, sheet_ref, header_cell_col, header_cell_row)
        if col_header not in existing_headers:
            new_files.append(f)

    if not new_files:
        print("No new files detected. Nothing to do.")
        return

    print(f"Found {len(source_files)} file(s) total, {len(new_files)} to process.")

    # Write row labels (only once, when output is brand-new)
    label_written = (ws.max_row > 1) or (not label_col)

    errors = []
    newly_added = 0

    for filepath in new_files:
        print(f"  Processing {filepath.name} ...", end=" ")
        try:
            values = read_range(filepath, sheet_ref,
                                col_start, row_start,
                                col_end, row_end)
        except Exception as exc:
            print(f"ERROR: {exc}")
            errors.append((filepath.name, str(exc)))
            continue

        col_header = filepath.stem if header_cell_row is None else read_header_cell(filepath, sheet_ref, header_cell_col, header_cell_row)

        target_col = next_col
        existing_headers[col_header] = target_col
        next_col += 1
        newly_added += 1

        # Write header
        hdr_cell = ws.cell(row=output_start_row, column=target_col, value=col_header)
        style_header(hdr_cell)

        # Write values starting one row below the header
        for i, val in enumerate(values):
            ws.cell(row=output_start_row + 1 + i, column=target_col, value=val)

        # Write row-label column from first file ever written
        if label_col and not label_written:
            label_col_idx = output_start_col - 1
            ws.cell(row=output_start_row, column=label_col_idx, value="Row Label")
            style_header(ws.cell(row=output_start_row, column=label_col_idx))
            for i in range(len(values)):
                lbl = ws.cell(row=output_start_row + 1 + i, column=label_col_idx,
                              value=f"Row {row_start + i}")
                style_label(lbl)
            label_written = True

        print("OK")

    # Column widths — only if specified
    if col_width is not None:
        for col in range(output_start_col, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = col_width

    wb_out.save(output_path)

    print()
    print(f"Done. {newly_added} new column(s) added.")
    if errors:
        print(f"\nWarning: {len(errors)} file(s) had errors and were skipped:")
        for name, msg in errors:
            print(f"  {name}: {msg}")
    print(f"Output: {output_path}")


# ---------------------------------------------------------------------------
# Interactive mode
# ---------------------------------------------------------------------------

def interactive():
    print("=" * 50)
    print("        Excel Combiner")
    print("=" * 50)
    print()

    # Input folder — loop until non-empty and valid
    while True:
        folder = clean_path(input("Path to the folder where your input files are: "))
        if not folder:
            print("  Please enter a path.")
            continue
        folder_path = Path(folder)
        if folder_path.is_dir():
            break
        print(f"  Error: '{folder}' is not a valid folder. Please try again.")

    # Output in same folder?
    while True:
        same = input("Should the output file be in the same folder? (yes/no): ").strip().lower()
        if same in ("yes", "y", "no", "n"):
            break
        print("  Please answer yes or no.")

    if same in ("yes", "y"):
        while True:
            filename = clean_path(input("Output filename (no extension needed, e.g. combined): "))
            if filename:
                stem = filename.split(".")[0]
                output_path = folder_path / (stem + ".xlsx")
                break
            print("  Please enter a filename.")
    else:
        while True:
            output_file = clean_path(input("Full path to the output file (no extension needed, e.g. C:/results/combined): "))
            if not output_file:
                print("  Please enter a path.")
                continue
            output_file_stem = Path(output_file).stem.split(".")[0]
            output_path = Path(output_file).parent / (output_file_stem + ".xlsx")
            if output_path.parent.exists() or output_path.parent == Path("."):
                break
            print(f"  Error: '{output_path.parent}' is not a valid folder. Please try again.")

    # Sheet — blank is valid (auto-detect)
    sheet = input("Sheet name or number (leave blank to auto-detect): ").strip()
    if sheet == "":
        sheet_ref = None
    else:
        try:
            sheet_ref = int(sheet)
        except ValueError:
            sheet_ref = sheet

    # Output start cell — where the first header goes
    while True:
        start_header = input("Where should the first header be in the output file? (leave blank for A1): ").strip()
        if not start_header:
            output_start_row, output_start_col = 1, 1
            break
        try:
            output_start_col, output_start_row = parse_cell(start_header)
            break
        except ValueError:
            print("  Error: invalid cell reference. Please try again (e.g. A1).")

    # Header source
    while True:
        header_choice = input("Use (1) filenames or (2) a specific cell in each file as column headers? (1/2): ").strip()
        if header_choice in ("1", "2"):
            break
        print("  Please enter 1 or 2.")

    header_cell_col, header_cell_row = None, None
    if header_choice == "2":
        while True:
            hdr_cell = input("Which cell contains the header in each input file? (e.g. A1): ").strip()
            if not hdr_cell:
                print("  Please enter a cell reference.")
                continue
            try:
                header_cell_col, header_cell_row = parse_cell(hdr_cell)
                break
            except ValueError:
                print("  Error: invalid cell reference. Please try again (e.g. A1).")

    # Start cell — loop only if empty or invalid
    while True:
        start_cell = input("Start cell (e.g. B3): ").strip()
        if not start_cell:
            print("  Please enter a cell reference.")
            continue
        try:
            col_start, row_start = parse_cell(start_cell)
            break
        except ValueError:
            print("  Error: invalid cell reference. Please try again (e.g. B3).")

    # End cell — loop only if empty, invalid, or different column
    while True:
        end_cell = input("End cell (e.g. B52): ").strip()
        if not end_cell:
            print("  Please enter a cell reference.")
            continue
        try:
            col_end, row_end = parse_cell(end_cell)
            if col_end != col_start:
                print("  Error: end cell must be in the same column as the start cell.")
            else:
                break
        except ValueError:
            print("  Error: invalid cell reference. Please try again (e.g. B52).")

    print()

    output_path.parent.mkdir(parents=True, exist_ok=True)

    build_or_update(
        folder=folder_path,
        output_path=output_path,
        sheet_ref=sheet_ref,
        col_start=col_start, row_start=row_start,
        col_end=col_end,     row_end=row_end,
        label_col=False,
        output_start_row=output_start_row,
        output_start_col=output_start_col,
        header_cell_col=header_cell_col,
        header_cell_row=header_cell_row,
        col_width=None,
    )

    input("\nPress Enter to exit.")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    # If no arguments passed, run interactive mode
    if len(sys.argv) == 1:
        interactive()
        return

    parser = argparse.ArgumentParser(
        description="Combine a cell range across many Excel files into one output file."
    )
    parser.add_argument("folder",      help="Folder containing source Excel files")
    parser.add_argument("output_file", help="Combined output .xlsx path")
    parser.add_argument("start_cell",  help="Top-left cell of range to read, e.g. B3")
    parser.add_argument("end_cell",    help="Bottom-right cell of range to read, e.g. B52")
    parser.add_argument("--sheet",     default=None,
                        help="Sheet name or 0-based index (optional)")
    parser.add_argument("--output-start-row", type=int, default=1,
                        help="Row in output where first header goes (default: 1)")
    parser.add_argument("--output-start-col", type=int, default=None,
                        help="Column in output where first header goes (default: 1, or 2 if --label-col)")
    parser.add_argument("--label-col", action="store_true",
                        help="Write row labels into the column before the first header")
    parser.add_argument("--col-width", type=float, default=None,
                        help="Set all column widths to this value (default: let Excel decide)")
    args = parser.parse_args()

    folder = Path(args.folder)
    if not folder.is_dir():
        sys.exit(f"Error: {folder} is not a directory.")

    output_path = Path(args.output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if args.sheet is None:
        sheet_ref = None
    else:
        try:
            sheet_ref = int(args.sheet)
        except ValueError:
            sheet_ref = args.sheet

    col_start, row_start = parse_cell(args.start_cell)
    col_end,   row_end   = parse_cell(args.end_cell)

    if col_start != col_end:
        sys.exit("Error: start and end cells must be in the same column "
                 "(this tool extracts one column per file).")

    if args.output_start_col is not None:
        output_start_col = args.output_start_col
    else:
        output_start_col = 2 if args.label_col else 1

    build_or_update(
        folder=folder,
        output_path=output_path,
        sheet_ref=sheet_ref,
        col_start=col_start, row_start=row_start,
        col_end=col_end,     row_end=row_end,
        label_col=args.label_col,
        output_start_row=args.output_start_row,
        output_start_col=output_start_col,
        col_width=args.col_width,
    )


if __name__ == "__main__":
    main()